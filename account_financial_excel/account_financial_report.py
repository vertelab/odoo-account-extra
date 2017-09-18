# -*- coding: utf-8 -*-
##############################################################################
#
# OpenERP, Open Source Management Solution, third party addon
# Copyright (C) 2004-2016  Vertel AB (<http://vertel.se>).
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

from openerp import api, models, fields, _

from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
import cStringIO
import base64

import logging
_logger = logging.getLogger(__name__)

class accounting_report(models.TransientModel):
    _inherit = "accounting.report"
    _description = "Accounting Report"

    state = fields.Selection(selection=[('get','get'),('put','put')],default=None)
    excel_file = fields.Binary()
    filename = fields.Char(default="financial_report.xls")

    @api.multi
    def check_excel(self):
            
        data = {}
        data['form'] = self.read(['date_from',  'date_to',  'fiscalyear_id', 'journal_ids', 'period_from', 'period_to',  'filter',  'chart_account_id', 'target_move','account_report_id', 'date_from_cmp',  'date_to_cmp',  'fiscalyear_id_cmp', 'journal_ids', 'period_from_cmp', 'period_to_cmp',  'filter_cmp',  'chart_account_id', 'target_move'])[0]
        data['form'] = self.read()[0]
        for field in ['fiscalyear_id','fiscalyear_id_cmp', 'chart_account_id', 'period_from_cmp', 'period_from','period_to_cmp','period_to', 'account_report_id']:
            if isinstance(data['form'][field], tuple):
                data['form'][field] = data['form'][field][0]
        comparison_context = self._build_comparison_context(data)
        data['form']['comparison_context'] = comparison_context
        used_context = self._build_contexts(data)
        data['form']['periods'] = used_context.get('periods', False) and used_context['periods'] or []
        data['form']['used_context'] = dict(used_context, lang=self._context.get('lang', 'en_US'))
        
        wb = Workbook()
        ws = wb.active
        
        level = [Font(bold=False,size=12),  # 0
               Font(bold=True,underline='singleAccounting',size=16),  # 1
               Font(bold=True,size=14), # 2
               Font(bold=True,size=12), # 3
               Font(bold=False,size=12), # 4
               Font(italic=True,size=10), # 5
               Font(size=10), # 6
               ]
        
        ws.title = self.account_report_id.name
        ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=6)
        ws.cell(row = 1, column = 1).value = ws.title
        r = 2

        ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=3)
        ws.cell(row = r, column = 1).value = _('Chart of Accounts:')
        ws.cell(row = r, column = 4).value = self.chart_account_id.name
        r += 1
        ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=3)
        ws.cell(row = r, column = 1).value = _('Fiscal Year:')
        ws.cell(row = r, column = 4).value = self.fiscalyear_id.name
        r += 1        
        ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=3)
        ws.cell(row = r, column = 1).value = _('Filter by:')
        ws.cell(row = r, column = 4).value = self.get_selection_value('filter',self.filter)
        if self.filter == 'filter_period':
            #~ raise Warning(self.period_from.name,self.period_to.name)
            ws.cell(row = r, column = 4).value = self.period_from.name
            ws.cell(row = r, column = 5).value = self.period_to.name
        if self.filter == 'filter_date':
            ws.cell(row = r, column = 4).value = self.date_from
            ws.cell(row = r, column = 5).value = self.date_to
        r += 1
        ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=3)
        ws.cell(row = r, column = 1).value = _('Target Moves:')
        ws.cell(row = r, column = 4).value = self.get_selection_value('target_move',self.target_move)
        r += 1


        if self.debit_credit:
            ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=5)
            ws.cell(row = r, column = 1).value = _('Names')
            ws.cell(row = r, column = 6).value = _('Debit')
            ws.cell(row = r, column = 7).value = _('Credit')
            ws.cell(row = r, column = 8).value = _('Balance')
            r += 1

            for line in self.get_lines(data):
                ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=5)
                ws.cell(row = r, column = 1).value = line['name']
                ws.cell(row = r, column = 1).font = level[line['level']]
                ws.cell(row = r, column = 6).value = line['debit']
                ws.cell(row = r, column = 6).font = level[line['level']]
                ws.cell(row = r, column = 7).value = line['credit']
                ws.cell(row = r, column = 7).font = level[line['level']]
                ws.cell(row = r, column = 8).value = line['balance']
                ws.cell(row = r, column = 8).font = level[line['level']]
                #~ ws.cell(row = r, column = 9).value = line['level']
                r += 1

        elif not self.enable_filter:
            ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=5)
            ws.cell(row = r, column = 1).value = _('Names')
            ws.cell(row = r, column = 6).value = _('Balance')
            r += 1
            for line in self.get_lines(data):
                ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=5)
                ws.cell(row = r, column = 1).value = line['name']
                ws.cell(row = r, column = 1).font = level[line['level']]
                ws.cell(row = r, column = 6).value = line['balance']
                ws.cell(row = r, column = 6).font = level[line['level']]
                #~ ws.cell(row = r, column = 9).value = line['level']

                r += 1
        else:
            ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=5)
            ws.cell(row = r, column = 1).value = _('Names')
            ws.cell(row = r, column = 6).value = _('Balance')
            ws.cell(row = r, column = 7).value = self.label_filter
            r += 1

            for line in self.get_lines(data):
                ws.merge_cells(start_row=r,end_row=r,start_column=1,end_column=5)
                ws.cell(row = r, column = 1).value = line['name']
                ws.cell(row = r, column = 1).font = level[line['level']]
                ws.cell(row = r, column = 6).value = line['balance']
                ws.cell(row = r, column = 6).font = level[line['level']]
                ws.cell(row = r, column = 7).value = line['balance_cmp']
                ws.cell(row = r, column = 7).font = level[line['level']]
                #~ ws.cell(row = r, column = 9).value = line['level']
                r += 1
        
        out = cStringIO.StringIO()
        wb.save(out)
        self.excel_file=base64.b64encode(out.getvalue())
        self.state = 'get'
        out.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'accounting.report',
            'view_mode': 'form',
            'view_type': 'form',
            'view_id': self.env.ref('account_financial_excel.accounting_report_excel_result').id,
            'res_id': self[0].id,
            'views': [(False, 'form')],
            'target': 'new',
        }

        return res
    
    # from account/report/account_financial_report.py    
    def get_lines(self, data):
        lines = []
        account_obj = self.pool.get('account.account')
        currency_obj = self.pool.get('res.currency')
        #~ ids2 = self.env['account.financial.report']._get_children_by_order([data['form']['account_report_id'][0]])
        ids2 = self.pool.get('account.financial.report')._get_children_by_order(self._cr, self._uid, [data['form']['account_report_id']], context=data['form']['used_context'])
        for report in self.pool.get('account.financial.report').browse(self._cr, self._uid, ids2, context=data['form']['used_context']):
            vals = {
                'name': report.name,
                'balance': report.balance * report.sign or 0.0,
                'type': 'report',
                'level': bool(report.style_overwrite) and report.style_overwrite or report.level,
                'account_type': report.type =='sum' and 'view' or False, #used to underline the financial report balances
            }
            if data['form']['debit_credit']:
                vals['debit'] = report.debit
                vals['credit'] = report.credit
            if data['form']['enable_filter']:
                vals['balance_cmp'] = self.pool.get('account.financial.report').browse(self._cr, self._uid, report.id, context=data['form']['comparison_context']).balance * report.sign or 0.0
            lines.append(vals)
            account_ids = []
            if report.display_detail == 'no_detail':
                #the rest of the loop is used to display the details of the financial report, so it's not needed here.
                continue
            if report.type == 'accounts' and report.account_ids:
                account_ids = account_obj._get_children_and_consol(self._cr, self._uid, [x.id for x in report.account_ids])
            elif report.type == 'account_type' and report.account_type_ids:
                account_ids = account_obj.search(self._cr, self._uid, [('user_type','in', [x.id for x in report.account_type_ids])])
            if account_ids:
                for account in account_obj.browse(self._cr, self._uid, account_ids, context=data['form']['used_context']):
                    #if there are accounts to display, we add them to the lines with a level equals to their level in
                    #the COA + 1 (to avoid having them with a too low level that would conflicts with the level of data
                    #financial reports for Assets, liabilities...)
                    if report.display_detail == 'detail_flat' and account.type == 'view':
                        continue
                    flag = False
                    vals = {
                        'name': account.code + ' ' + account.name,
                        'balance':  account.balance != 0 and account.balance * report.sign or account.balance,
                        'type': 'account',
                        'level': report.display_detail == 'detail_with_hierarchy' and min(account.level + 1,6) or 6, #account.level + 1
                        'account_type': account.type,
                    }

                    if data['form']['debit_credit']:
                        vals['debit'] = account.debit
                        vals['credit'] = account.credit
                    if not currency_obj.is_zero(self._cr, self._uid, account.company_id.currency_id, vals['balance']):
                        flag = True
                    if data['form']['enable_filter']:
                        vals['balance_cmp'] = account_obj.browse(self._cr, self._uid, account.id, context=data['form']['comparison_context']).balance * report.sign or 0.0
                        if not currency_obj.is_zero(self._cr, self._uid, account.company_id.currency_id, vals['balance_cmp']):
                            flag = True
                    if flag:
                        lines.append(vals)
        return lines

    def get_selection_text(self,field,value):
        for type,text in self.fields_get([field])[field]['selection']:
                if text == value:
                    return type
        return None

    def get_selection_value(self,field,value):
        for type,text in self.fields_get([field])[field]['selection']:
                if type == value:
                    return text
        return None
     
# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
